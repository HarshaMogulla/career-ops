#!/usr/bin/env python3
"""
Export applications.md + pipeline.md → tracker.xlsx

Two sheets:
  - Applications: all evaluated / applied roles with score color-coding
  - Pipeline:     URLs from scan inbox (pending/processed)

Source of truth stays in the .md files — this is a READ-ONLY export for readability.
Run after any tracker update: `python3 export_tracker.py` or `npm run tracker`.
"""
import re
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).parent
APPLICATIONS_MD = ROOT / "data" / "applications.md"
PIPELINE_MD = ROOT / "data" / "pipeline.md"
OUT_XLSX = ROOT / "data" / "tracker.xlsx"

SCORE_COLORS = {
    "great":   "C6EFCE",  # green for 4.0+
    "ok":      "FFEB9C",  # yellow for 3.0-4.0
    "poor":    "FFC7CE",  # red for <3.0
    "skip":    "D9D9D9",  # gray for SKIP/Rejected
}
HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
THIN = Side(style="thin", color="B0B0B0")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def parse_applications_table(text):
    rows = []
    for line in text.splitlines():
        line = line.strip()
        if not line.startswith("|") or "|---" in line:
            continue
        cells = [c.strip() for c in line.strip("|").split("|")]
        if not cells or cells[0].lower() == "#":
            continue
        if not cells[0].isdigit():
            continue
        while len(cells) < 9:
            cells.append("")
        rows.append(cells[:9])
    return rows


def parse_pipeline(text):
    """Return list of parsed pipeline rows.

    Handles three line formats:
      - [!] URL | Company | Role — Error: description      → Error
      - [ ] URL | Company | Role                           → Pending
      - [x] #NNN | URL | Company | Role | X.X/5 | PDF ✅   → Processed
    """
    rows = []
    section_hint = "Pending"
    for line in text.splitlines():
        line = line.strip()
        if line.startswith("## Pendientes") or line.lower().startswith("## pending"):
            section_hint = "Pending"
            continue
        if line.startswith("## Procesadas") or line.lower().startswith("## processed"):
            section_hint = "Processed"
            continue
        m = re.match(r"-\s*\[([ x!])\]\s*(.+)", line)
        if not m:
            continue
        marker, body = m.group(1), m.group(2)

        # Determine status from marker (takes precedence over section heading)
        if marker == "!":
            status = "Error"
        elif marker == "x":
            status = "Processed"
        else:
            status = section_hint

        # Default row fields
        row = {
            "status": status, "num": "", "url": "", "company": "",
            "role": "", "score": "", "pdf": "", "notes": "",
        }

        # Split on " — " first to isolate error description
        error_note = ""
        if " — Error:" in body:
            body, _, err = body.partition(" — Error:")
            error_note = err.strip()

        parts = [p.strip() for p in body.split("|")]

        # Processed format: "#NNN | URL | Company | Role | X.X/5 | PDF ✅"
        if parts and parts[0].startswith("#"):
            row["num"] = parts[0].lstrip("#").strip()
            if len(parts) > 1: row["url"] = parts[1]
            if len(parts) > 2: row["company"] = parts[2]
            if len(parts) > 3: row["role"] = parts[3]
            if len(parts) > 4: row["score"] = parts[4]
            if len(parts) > 5:
                pdf_part = parts[5]
                if "✅" in pdf_part: row["pdf"] = "✅"
                elif "❌" in pdf_part: row["pdf"] = "❌"
                else: row["pdf"] = pdf_part
            if len(parts) > 6: row["notes"] = " · ".join(parts[6:])
        else:
            # Pending / Error format: "URL | Company | Role"
            if len(parts) > 0: row["url"] = parts[0]
            if len(parts) > 1: row["company"] = parts[1]
            if len(parts) > 2: row["role"] = parts[2]

        if error_note:
            row["notes"] = (row["notes"] + " · " if row["notes"] else "") + f"Error: {error_note}"

        rows.append(row)
    return rows


def score_fill(score_text):
    if not score_text:
        return None
    if "SKIP" in score_text.upper() or "REJECT" in score_text.upper():
        return PatternFill("solid", fgColor=SCORE_COLORS["skip"])
    m = re.search(r"(\d+\.?\d*)\s*/\s*5", score_text)
    if not m:
        return None
    v = float(m.group(1))
    if v >= 4.0:
        return PatternFill("solid", fgColor=SCORE_COLORS["great"])
    if v >= 3.0:
        return PatternFill("solid", fgColor=SCORE_COLORS["ok"])
    return PatternFill("solid", fgColor=SCORE_COLORS["poor"])


def write_applications(ws, rows):
    headers = ["#", "Date", "Company", "Role", "Score", "Status", "PDF", "Report", "Notes"]
    ws.append(headers)
    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="left", vertical="center")
        cell.border = BORDER
    for r in rows:
        ws.append(r)
        row_idx = ws.max_row
        # Color score cell
        score_cell = ws.cell(row=row_idx, column=5)
        fill = score_fill(f"{r[4]} {r[5]}")
        if fill:
            score_cell.fill = fill
        # Borders + alignment
        for col_idx in range(1, 10):
            c = ws.cell(row=row_idx, column=col_idx)
            c.border = BORDER
            c.alignment = Alignment(vertical="top", wrap_text=True)
    # Column widths
    widths = [5, 12, 22, 40, 8, 12, 5, 40, 60]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    # Freeze header
    ws.freeze_panes = "A2"


def write_pipeline(ws, rows):
    headers = ["Status", "#", "Company", "Role", "Score", "PDF", "URL", "Notes"]
    ws.append(headers)
    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="left", vertical="center")
        cell.border = BORDER

    status_fills = {
        "Processed": PatternFill("solid", fgColor="C6EFCE"),    # green
        "Pending":   PatternFill("solid", fgColor="FFF2CC"),    # yellow
        "Error":     PatternFill("solid", fgColor="FFC7CE"),    # red
    }

    # Sort: Processed first (most recent work), then Pending, then Error
    order = {"Processed": 0, "Pending": 1, "Error": 2}
    rows = sorted(rows, key=lambda r: (order.get(r["status"], 99), r.get("num") or ""))

    for r in rows:
        ws.append([
            r["status"],
            r["num"],
            r["company"],
            r["role"],
            r["score"],
            r["pdf"],
            r["url"],
            r["notes"],
        ])
        row_idx = ws.max_row
        # Status cell colored by status
        status_cell = ws.cell(row=row_idx, column=1)
        if r["status"] in status_fills:
            status_cell.fill = status_fills[r["status"]]
        # Score cell colored by score value
        sf = score_fill(r["score"])
        if sf:
            ws.cell(row=row_idx, column=5).fill = sf
        # Borders + wrap
        for col_idx in range(1, 9):
            c = ws.cell(row=row_idx, column=col_idx)
            c.border = BORDER
            c.alignment = Alignment(vertical="top", wrap_text=True)

    widths = [10, 5, 22, 38, 10, 5, 65, 45]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"


def main():
    wb = Workbook()
    ws_apps = wb.active
    ws_apps.title = "Applications"
    ws_pipe = wb.create_sheet("Pipeline")

    apps_text = APPLICATIONS_MD.read_text() if APPLICATIONS_MD.exists() else ""
    pipe_text = PIPELINE_MD.read_text() if PIPELINE_MD.exists() else ""

    app_rows = parse_applications_table(apps_text)
    pipe_rows = parse_pipeline(pipe_text)

    write_applications(ws_apps, app_rows)
    write_pipeline(ws_pipe, pipe_rows)

    wb.save(str(OUT_XLSX))
    print(f"✅ tracker.xlsx written: {OUT_XLSX}")
    print(f"   Applications: {len(app_rows)} rows")
    print(f"   Pipeline:     {len(pipe_rows)} rows")


if __name__ == "__main__":
    main()
